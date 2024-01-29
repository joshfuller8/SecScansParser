import openpyxl


def count_unique_entries(source_file_path, customer_nums_to_ignore):
    unique_entries = []

    # Load the workbook
    workbook = openpyxl.load_workbook(source_file_path, data_only=True)
    scan_count = 0
    # Iterate through sheets in the workbook
    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith("TABLET"):
            sheet = workbook[sheet_name]

            # Iterate through column A (assuming column A is column 1)
            for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True):
                for cell_value in row:
                    if not isinstance(cell_value, str):
                        continue
                    if cell_value and cell_value.startswith("CN"):
                        if cell_value not in customer_nums_to_ignore:
                            unique_entries.append(cell_value)
                            scan_count += 1
    values = [len(set(unique_entries)), scan_count]
    print(" ")
    print(f"Head Count: {values[0]}")
    print(f"Scan Count: {values[1]}")
    return values


# source_file_path = "C:\\Users\\Josh\\Desktop\\scans.xlsx"
# values = count_unique_entries(workbook_path)
# print(" ")
# print(f"Head Count: {values[0]}")
# print(f"Scan Count: {values[1]}")
