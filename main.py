import openpyxl
from headcount import count_unique_entries
from Pull_Validated_Data import pull_validated_data
from filter_scans_by_security_triggers import create_list_of_filtered_scans

# Source Excel file path
source_file_path = "C:\\Users\\Josh\\Desktop\\RRI SecScans Parser\\book 8.xlsx"

# filepath and name of the new file that will be created
new_file_path = "C:\\Users\\Josh\\Desktop\\RRI SecScans Parser\\filtered_scans.xlsx"

# Load the source Excel file
workbook = openpyxl.load_workbook(source_file_path, data_only=True)

# Time thresholds (in seconds) for calculating time_diff
trigger_parameters = {
    'SAME_TABLET_LOWER': 5,
    'SAME_TABLET_UPPER': 300,
    'DIFF_TABLET_UPPER': 300
    }

# turn triggers on or off, True is on
triggers = {
    'same_tablet_within_range': True,
    'diff_tablet_under_5_mins': True
}

# any CN in this list will not be counted, all CN from Crew sheet will be automatically added to this list
customer_nums_to_ignore = [
    'CN-0003544198'
]

# pull CN#s of crew at event to ignore in counts and add to customer_nums_to_ignore
crew_worksheet = workbook['Crew']
for i in range(1, crew_worksheet.max_row + 1):
    customer_nums_to_ignore.append(crew_worksheet[f'A{i}'].value)

print(customer_nums_to_ignore)
validated_scans_by_cn = pull_validated_data(workbook)  # creates a dictionary of validated data in format
# {CN:[scan,scan,...], CN:[scan,scan,....)} where the keys are each unique CN that has been scanned and the scans are
# each of the scans that correspond to that CN
# function is located in Pull_Validated_Data.py
# print(validated_scans_by_cn)
print('inputs validated')
print('filtering scans')

filtered_rows_that_meet_criteria = create_list_of_filtered_scans(validated_scans_by_cn, customer_nums_to_ignore,
                                                                 triggers, trigger_parameters)  # parses the validated
# scans and returns a new dictionary of only entries that meet the criteria for the security triggers. They can be
# configured at the top of main. triggers is to turn on or off different security triggers and trigger_parameters
# contains the data needed for those triggers.

if filtered_rows_that_meet_criteria:
    print('scans filtered')
else:
    print('no results')

values = count_unique_entries(source_file_path, customer_nums_to_ignore)  # counts all validated scans and unique scans
# to return a headcount and a total of all scans

# Create a new workbook and worksheet to add the filtered rows to
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

# Write headers
headers = ["CN#", "Timestamp", "Tablet", "Location", "Headcount", "Total Scans"]
new_sheet.append(headers)

# Write filtered rows
for row_data in filtered_rows_that_meet_criteria:
    new_sheet.append(row_data)
new_sheet['E2'] = values[0]
new_sheet['F2'] = values[1]

# format column widths
column_letters = ["A", "B", "C", "D", "E", "F"]
# data labels      CN#   Datetime  Tablet Location  heads   scans
column_widths = {"A": 15, "B": 19, "C": 6, "D": 8, "E": 10, "F": 10}
index = 0
for header in headers:
    new_sheet.column_dimensions[column_letters[index]].width = column_widths[column_letters[index]]
    index += 1

# Save the new workbook
new_workbook.save(new_file_path)
